import os
import shutil
from datetime import datetime, timedelta
import schedule
import time
import oracledb
from dotenv import load_dotenv
import argparse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from copy import copy


class EnvConfig:
    def __init__(self, args=None):
        """初始化配置参数（命令行参数优先于环境变量）"""
        self._load_config(args)

    def _load_config(self, args) -> None:
        """加载配置参数的内部方法"""
        config_mapping = {
            # 数据库配置
            'host': ('DB_HOST', 'localhost'),
            'user': ('DB_USER', 'system'),
            'password': ('DB_PASSWORD', ''),
            'service_name': ('DB_NAME', 'orcl'),
            'port': ('DB_PORT', 1521),
            'encoding': ('DB_ENCODING', 'UTF-8'),
            # Excel配置
            'excel_path': ('EXCEL_PATH', 'report.xlsx'),
            'sheet_name': ('SHEET_NAME', 'Sheet1'),
            'schedule_time': ('SCHEDULE_TIME', '09:00'),
            # 列名配置
            'col_start_time': ('COL_START_TIME', '开始时间'),
            'col_end_time': ('COL_END_TIME', '结束时间'),
            'col_guest_count': ('COL_GUEST_COUNT', '入住旅客数'),
            'col_late_upload': ('COL_LATE_UPLOAD', '15分上传不及时数'),  # 修正参数名保持一致
            'col_completion_rate': ('COL_COMPLETION_RATE', '完成率'),
            # 时间参数
            'start_time': ('START_TIME', ''),
            'end_time': ('END_TIME', '')
        }

        for attr, (env_var, default) in config_mapping.items():
            # 类型转换函数(根据属性名特殊处理)
            converter = lambda x: int(x) if attr == 'port' else x
            # 获取参数值（命令行参数优先）
            arg_value = getattr(args, attr, None) if args else None
            env_value = os.getenv(env_var, default)
            
            # 特殊处理端口号的类型转换
            if attr == 'port' and not isinstance(env_value, int):
                env_value = int(env_value) if env_value else default
            
            setattr(self, attr, arg_value or env_value)


class ExcelProcessor:
    def __init__(self, config):
        self.config = config
        self.excel_path = config.excel_path
        self.sheet_name = config.sheet_name

    def backup_excel(self):
        """备份Excel文件"""
        backup_path = f"{self.excel_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shutil.copy2(self.excel_path, backup_path)
        return backup_path

    def update_excel(self, start_time: str, end_time: str, count1: int, count2: int) -> None:
        """更新Excel文件并保持原有格式
        
        Args:
            start_time: 开始时间字符串（格式：YYYYMMDDHHMISS）
            end_time: 结束时间字符串（格式：YYYYMMDDHHMISS）
            count1: 入住旅客数
            count2: 15分上传不及时数
        """
        try:
            # 加载工作簿并获取目标工作表
            workbook = openpyxl.load_workbook(self.excel_path, data_only=False)
            sheet = workbook[self.sheet_name]

            # 获取列头映射（使用字典推导式优化）
            headers = {
                cell.value: col
                for col in range(1, sheet.max_column + 1)
                if (cell := sheet.cell(row=1, column=col)).value in {
                    self.config.col_start_time,
                    self.config.col_end_time,
                    self.config.col_guest_count,
                    self.config.col_late_upload,
                    self.config.col_completion_rate
                }
            }

            # 计算插入位置（排除最后的合计行）
            last_data_row = sheet.max_row - 1

            # 插入新行在倒数第二行（保持合计行在最后）
            sheet.insert_rows(last_data_row + 1)
            new_row_num = last_data_row + 1

            # 复制上一行的格式到新行（优化样式复制）
            if last_data_row >= 1:  # 确保至少有一行数据
                source_row = sheet[last_data_row]
                target_row = sheet[new_row_num]
                for src_cell, tgt_cell in zip(source_row, target_row):
                    if src_cell.has_style:
                        # 复制完整单元格样式
                        tgt_cell._style = copy(src_cell._style)
                        # 显式复制关键格式属性
                        tgt_cell.number_format = src_cell.number_format
                        tgt_cell.font = copy(src_cell.font)
                        tgt_cell.border = copy(src_cell.border)
                        tgt_cell.fill = copy(src_cell.fill)
                        tgt_cell.alignment = copy(src_cell.alignment)
                        # 复制列宽设置（增加异常处理）
                        src_col = get_column_letter(src_cell.column)
                        tgt_col = get_column_letter(tgt_cell.column)
                        try:
                            # 安全获取列宽设置
                            src_width = sheet.column_dimensions[src_col].width
                            sheet.column_dimensions[tgt_col].width = src_width
                        except KeyError:
                            # 源列不存在时使用默认列宽并记录警告
                            sheet.column_dimensions[tgt_col].width = 15
                            print(f"警告：列 {src_col} 的宽度未定义，已使用默认值 15")
                        tgt_cell.font = copy(src_cell.font)
                        tgt_cell.border = copy(src_cell.border)
                        tgt_cell.fill = copy(src_cell.fill)
                        tgt_cell.alignment = copy(src_cell.alignment)

            # 写入新数据并设置适当的格式
            col_data = {
                self.config.col_start_time: (start_time, 'date'),
                self.config.col_end_time: (end_time, 'date'),
                self.config.col_guest_count: (count1, 'number'),
                self.config.col_late_upload: (count2, 'number')
            }

            for header, (value, data_type) in col_data.items():
                if header in headers:
                    cell = sheet.cell(row=new_row_num, column=headers[header])
                    # 确保时间值作为字符串写入，防止Excel自动合并
                    if data_type == 'date':
                        cell.value = str(value)
                        cell.number_format = '@'  # 设置为文本格式
                    else:
                        cell.value = value
                        if data_type == 'number':
                            cell.number_format = '0'

            # 更新完成率列的公式
            if self.config.col_completion_rate in headers:
                completion_col = headers[self.config.col_completion_rate]
                count1_col = headers[self.config.col_guest_count]
                count2_col = headers[self.config.col_late_upload]
                
                # 获取完成率的计算公式
                formula_cell = sheet.cell(row=last_row, column=completion_col)
                formula_value = str(formula_cell.value)
                if formula_value.startswith('='):
                    # 如果存在公式，复制并更新行号
                    new_formula = formula_value.replace(str(last_row), str(new_row_num))
                    sheet.cell(row=new_row_num, column=completion_col).value = new_formula
                else:
                    # 如果没有公式，创建新的公式
                    count1_letter = openpyxl.utils.get_column_letter(count1_col)
                    count2_letter = openpyxl.utils.get_column_letter(count2_col)
                    formula = f"={count1_letter}{new_row_num}/({count1_letter}{new_row_num}+{count2_letter}{new_row_num})"
                    sheet.cell(row=new_row_num, column=completion_col).value = formula
                    sheet.cell(row=new_row_num, column=completion_col).number_format = '0.00%'

            # 更新合计行的公式范围
            total_row = sheet.max_row
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=total_row, column=col)
                cell_value = str(cell.value)
                if cell_value.startswith('='):
                    # 替换SUM公式范围
                    if 'SUM' in cell_value.upper():
                        start_ref = cell_value[cell_value.find('(') + 1:cell_value.find(':')]
                        # 获取当前列号
                        col_letter = get_column_letter(cell.column)
                        end_ref = f"{col_letter}{total_row-1}"
                        new_formula = f"=SUM({start_ref}:{end_ref})"
                    # 最后一列的完成率公式
                    elif col == completion_col:
                        count1_letter = openpyxl.utils.get_column_letter(col-2)
                        count2_letter = openpyxl.utils.get_column_letter(col-1)
                        new_formula = f"={count1_letter}{total_row}/({count1_letter}{total_row}+{count2_letter}{total_row})"
                        cell.value = new_formula
                        cell.number_format = '0.00%'
            # 保存工作簿
            workbook.save(self.excel_path)
            print(f"已更新Excel文件 {self.excel_path}")

        except Exception as e:
            print(f"更新Excel时发生错误: {str(e)}")
            raise


class DatabaseQuery:
    # 定义SQL查询语句为类常量
    # 基础查询：统计时间段内有效入住记录
    BASE_QUERY = """
        SELECT COUNT(1) 
        FROM v_gnlkxx_50 v
        WHERE v.rzsj >= TO_DATE(:start_time, 'YYYYMMDDHH24MISS')
          AND v.rzsj < TO_DATE(:end_time, 'YYYYMMDDHH24MISS')
    """
    
    LATE_UPLOAD_CONDITION = """
        AND ((v.tfsj IS NULL AND (v.gxsj - v.rzsj) * 24 * 60 >= 15)
        OR (v.tfsj IS NOT NULL AND (v.gxsj - v.tfsj) * 24 * 60 >= 15))
    """

    def __init__(self, config: EnvConfig):
        """初始化数据库查询器
        Args:
            config: 包含数据库配置的EnvConfig实例
        """
        self.config = config
        # 修正连接字符串格式（SID使用//host:port/SID，Service Name使用host:port/ServiceName）
        # 初始化Oracle客户端库（添加lib_dir参数如果需要）
        oracledb.init_oracle_client()
        # 创建带服务名的DSN
        self.dsn = oracledb.makedsn(
            config.host,
            config.port,
            service_name=config.service_name
        )

    def execute_queries(self, start_time: str, end_time: str) -> tuple[int, int]:
        """执行并返回两个查询结果
        Args:
            start_time: 开始时间（YYYYMMDDHH24MISS格式）
            end_time: 结束时间（YYYYMMDDHH24MISS格式）
        Returns:
            元组包含 (普通查询结果, 延迟上传查询结果)
        """
        try:
            with oracledb.connect(
                user=self.config.user,
                password=self.config.password,
                dsn=self.dsn
            ) as connection:
                with connection.cursor() as cursor:
                    # 执行基础查询
                    cursor.execute(self.BASE_QUERY,
                                 start_time=start_time,
                                 end_time=end_time)
                    count1 = cursor.fetchone()[0]

                    # 执行带延迟条件的查询
                    cursor.execute(f"{self.BASE_QUERY} {self.LATE_UPLOAD_CONDITION}",
                                 start_time=start_time,
                                 end_time=end_time)
                    count2 = cursor.fetchone()[0]

                    return count1, count2

        except Exception as e:
            print(f"执行数据库查询时发生错误: {str(e)}")
            raise


def job(config):
    """定时任务主函数"""
    try:
        # 设置时间范围
        # 优先使用配置的时间参数，如果没有配置则使用默认时间范围
        if config.start_time and config.end_time:
            # 转换时间格式
            try:
                # 尝试解析为YYYY-MM-DD HH:MM:SS格式
                start_dt = datetime.strptime(config.start_time, '%Y-%m-%d %H:%M:%S')
                end_dt = datetime.strptime(config.end_time, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    # 尝试解析为YYYY-MM-DD格式
                    start_dt = datetime.strptime(config.start_time, '%Y-%m-%d')
                    end_dt = datetime.strptime(config.end_time, '%Y-%m-%d')
                except ValueError:
                    # 尝试解析为YYYYMMDDHHMISS格式
                    start_dt = datetime.strptime(config.start_time, '%Y%m%d%H%M%S')
                    end_dt = datetime.strptime(config.end_time, '%Y%m%d%H%M%S')
            
            # 转换为数据库需要的格式
            start_time = start_dt.strftime('%Y%m%d%H%M%S')
            print(f"转换后的开始时间：{start_time}")
            end_time = end_dt.strftime('%Y%m%d%H%M%S')
        else:
            end_time = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            start_time = end_time - timedelta(days=1)
        
        # 初始化处理器
        db_query = DatabaseQuery(config)
        excel_processor = ExcelProcessor(config)
        
        # 备份Excel
        backup_path = excel_processor.backup_excel()
        print(f"Excel已备份到: {backup_path}")
        
        # 使用已格式化的时间字符串执行查询和更新
        count1, count2 = db_query.execute_queries(start_time, end_time)
        
        # 更新Excel(直接使用已格式化时间字符串)
        excel_processor.update_excel(start_time, end_time, count1, count2)
        print(f"Excel更新成功 | 时间范围: {start_time} 至 {end_time}")
        
    except Exception as e:
        print(f"执行任务时发生错误: {str(e)}")


def main():
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='SQL查询结果导出到Excel工具')
    
    # 添加数据库连接参数
    parser.add_argument('--host', help='数据库主机地址')
    parser.add_argument('--port', type=int, help='数据库端口')
    parser.add_argument('--user', help='数据库用户名')
    parser.add_argument('--password', help='数据库密码')
    parser.add_argument('--dbname', help='数据库名称')
    parser.add_argument('--encoding', help='数据库编码')
    
    # 添加Excel相关参数
    parser.add_argument('--excel-path', help='Excel文件路径')
    parser.add_argument('--sheet-name', help='Excel工作表名称')
    
    # 添加Excel列名参数
    parser.add_argument('--col-start-time', help='开始时间列名')
    parser.add_argument('--col-end-time', help='结束时间列名')
    parser.add_argument('--col-guest-count', help='入住旅客数列名')
    parser.add_argument('--col-late-upload', help='15分上传不及时数列名')
    parser.add_argument('--col-completion-rate', help='完成率列名')
    
    # 添加时间参数
    parser.add_argument('--start-time', help='开始时间，支持格式：YYYY-MM-DD HH:MM:SS 或 YYYY-MM-DD 或 YYYYMMDDHHMISS')
    parser.add_argument('--end-time', help='结束时间，支持格式：YYYY-MM-DD HH:MM:SS 或 YYYY-MM-DD 或 YYYYMMDDHHMISS')
    
    # 添加运行模式参数
    parser.add_argument('--run-now', action='store_true', help='立即执行一次，不等待定时任务')
    parser.add_argument('--config', help='指定配置文件路径')

    # 添加调度参数
    parser.add_argument('--schedule-time', help='定时任务时间，格式为 HH:MM')
    
    args = parser.parse_args()
    
    # 如果指定了配置文件，加载配置文件
    if args.config:
        load_dotenv(args.config)
    else:
        load_dotenv()

     
    # 创建配置对象
    config = EnvConfig(args)
    
    # 初始化Oracle thick mode
    oracledb.init_oracle_client()
    # 根据参数决定是立即执行还是运行定时任务
    if args.run_now:
        job(config)
    else:
        print("每天{}点执行任务".format(config.schedule_time))
        # 设置定时任务
        schedule.every().day.at(config.schedule_time).do(job, config)
        
        while True:
            schedule.run_pending()
            time.sleep(60)


if __name__ == "__main__":
    main()
