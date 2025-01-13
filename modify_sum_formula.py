import openpyxl
from openpyxl.utils import get_column_letter
from shutil import copyfile

# 备份文件
src_file = '2024年部级上报数据汇总.xlsx'
backup_file = '2024年部级上报数据汇总.xlsx.backup_20250113_182109'
copyfile(src_file, backup_file)

# 打开Excel文件
wb = openpyxl.load_workbook(src_file)
ws = wb.active

# 查找包含SUM公式的单元格
for row in ws.iter_rows():
    for cell in row:
        if cell.data_type == 'f' and 'SUM' in cell.value:
            # 获取当前列号
            col = get_column_letter(cell.column)
            
            # 创建动态范围公式
            start_row = 2  # 假设数据从第2行开始
            end_row = ws.max_row-1
            new_formula = f'=SUM({col}{start_row}:{col}{end_row})'
            
            # 更新公式
            cell.value = new_formula

# 保存修改
wb.save(src_file)
print(f"SUM公式已更新，原始文件已备份为：{backup_file}")